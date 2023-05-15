from openpyxl import Workbook, load_workbook
from datetime import datetime
#-------------------------------------------------------------
#SETUP
#-------------------------------------------------------------
wb1 = load_workbook(‘MasterPriceList.xlsx')
wb2 = load_workbook(‘TemplatePriceList.xlsx')

ws1 = wb1.active
ws2 = wb2.active
#-------------------------------------------------------------
#FUNCTIONS
#-------------------------------------------------------------
def ingestToDict(ws):
    sheet_cells = {}
    rowNum = 2
    for row in ws.iter_rows(min_row=2):
        #row 0 is the first cell in the master "product code"
        pCode = row[0].value
        sheet_cells[pCode] = rowNum
        rowNum += 1
    return sheet_cells


def createRow(row):
    row_cells = []
    for cell in range(0, len(row)):
        row_cells.append(row[cell].value)
    now = datetime.now()
    row_cells.append(now)
    return row_cells


def ingestToList(ws):
    sheet_cells = []
    for row in ws.iter_rows(min_row=2):
        newRow = createRow(row)
        sheet_cells.append(newRow)
    return sheet_cells


def mergeSheets(master, template):
    matches = 0
    noMatch = 0
    for row in template:
        #row 0 is the first cell in the template "product code"
        if row[0] in master:
            matches += 1
            masterIdx = master[row[0]]
            #column 1 is the first cell in the master "product code"
            #row 0 is the first cell in the template "product code"
            ws1.cell(row=masterIdx, column=9, value=row[8])
            now = datetime.now()
            ws1.cell(row=masterIdx, column=11, value=now)
        else:
            noMatch += 1

    print(f"Matches: {matches}, No match: {noMatch}")
#-------------------------------------------------------------
#IMPLEMENTATION
#-------------------------------------------------------------
templateList = ingestToList(ws2)
masterDict = ingestToDict(ws1)
mergeSheets(masterDict, templateList)
#-------------------------------------------------------------
#CLOSING
#-------------------------------------------------------------
wb1.save(‘MasterPriceList.xlsx')
wb2.save(‘TemplatePriceList.xlsx')